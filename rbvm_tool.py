# -*- coding: utf-8 -*-

## ==== IMPORTS  ==== ##

# Imports standards
import sys
import os
import re
import json
import math
import time
import secrets
import string
import sqlite3
import requests
from datetime import datetime
from contextlib import ExitStack
import numpy as np

# D√©sactiver l'avertissement indiquant que certaines fonctionnalit√©s du fichier excel ne sont pas prise en charge comme "Conditional Formatting extension is not supported and will be removed" et "Data Validation extension is not supported and will be removed" 
import warnings
warnings.simplefilter("ignore", UserWarning)


# Imports pour la gestion des fichiers et des donn√©es
import openpyxl  # Manipulation des fichiers Excel
import pandas as pd  # Manipulation et analyse de donn√©es sous forme de DataFrames
from tabulate import tabulate  # Formatage des tableaux pour l'affichage CLI

# Imports pour l'affichage graphique
import matplotlib.pyplot as plt  # G√©n√©ration de graphiques
import matplotlib.image as mpimg
import matplotlib.patches as mpatches
from io import BytesIO

# Imports PyQt5 pour l'interface utilisateur
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QLineEdit,
    QPushButton,
    QFileDialog,
    QVBoxLayout,
    QWidget,
    QGroupBox,
    QHBoxLayout,
    QProgressBar,
    QRadioButton,
    QMessageBox,
    QInputDialog,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer

# Imports Tkinter pour les bo√Ætes de dialogue de fichiers
from tkinter import Tk, filedialog

##### VARIABLES GLOBALES #####
attack_vector = None
attack_complexity = None
privileges_required = None
user_interaction = None
scope = None
confidentiality = None
integrity = None
availability = None
authentification = None
exp_score = None
env_score = None
kev = None
folder_BS_VERT = None
folder_BS_ORANGE = None
folder_BS_ROUGE = None
folder_VM_VERT = None
folder_VM_ORANGE = None
folder_VM_ROUGE = None
folder_VM_META = None

# Dictionnaire permettant de stocker les biens supports ainsi que leurs valeurs P1, P2, P3, P4, P5,
dicoListePx = {}
dicoC = {"p5": [], "p4": [], "p3": [], "p2": [], "p1": []}
dicoI = {"p5": [], "p4": [], "p3": [], "p2": [], "p1": []}
dicoA = {"p5": [], "p4": [], "p3": [], "p2": [], "p1": []}
dicoGlobal = {"confidentiality": dicoC, "integrity": dicoI, "availability": dicoA}


# Dictionnaire de conversion CVSS 3.1 (lettre -> valeur num√©rique)
CVSS_NUMERIC_VALUES = {
    "AV": {"N": 0.85, "A": 0.62, "L": 0.55, "P": 0.2},
    "AC": {"L": 0.77, "H": 0.44},
    "PR": {"N": 0.85, "L": 0.62, "H": 0.27},
    "UI": {"N": 0.85, "R": 0.62},
    "S": {"U": 0, "C": 1},
    "C": {"H": 0.56, "L": 0.22, "N": 0},
    "I": {"H": 0.56, "L": 0.22, "N": 0},
    "A": {"H": 0.56, "L": 0.22, "N": 0}
}

## ==== PARTIE N¬∞1 D√©finition de l'IHM ==== ##

class RBVMTool(QMainWindow):
    def __init__(self):
        super().__init__()

        self.passphrase = self.get_secure_passphrase()
        global passphrase
        passphrase = self.passphrase
        print("\nPassphrase saisie avec succ√®s")

        # Initialisation de la connexion SQLite dans l'objet
        self.conn = sqlite3.connect("rbvm.db")
        self.cur = self.conn.cursor()

        self.setWindowTitle("Risk Based Vulnerability Management (RBVM) Tool")
        self.setGeometry(200, 100, 700, 800)
        self.vdr_data_list = []  # Stocker les VDR s√©lectionn√©s
        self.kev_data = {}  # Stocker les donn√©es KEV
        self.initUI()

# Famille de fonction pour la passphrase
    def get_secure_passphrase(self):
        while True:
            passphrase, ok = self.ask_for_passphrase()

            if not ok:
                print("\nOp√©ration annul√©e")
                sys.exit()

            if self.is_passphrase_valid(passphrase):
                print("\nPassphrase saisie avec succ√®s")
                return passphrase
            else:
                QMessageBox.warning(
                    None,
                    "Passphrase invalide",
                    "Votre passphrase ne respecte pas les exigences !\n\n"
                    "Elle doit contenir au moins :\n"
                    "- 8 caract√®res\n"
                    "- 1 majuscule\n"
                    "- 1 chiffre\n"
                    "- 1 caract√®re sp√©cial",
                ) # Demande une passphrase avec option de g√©n√©ration automatique
    def ask_for_passphrase(self):
        dialog = QMessageBox()
        dialog.setWindowTitle("Chiffrement BDD obligatoire")
        dialog.setText(
            "Veuillez saisir une passphrase forte ou en g√©n√©rer une automatiquement."
        )

        passphrase_input = QLineEdit()
        passphrase_input.setEchoMode(QLineEdit.Password)

        generate_button = QPushButton("G√©n√©rer une passphrase forte")
        copy_button = QPushButton("Copier")

        layout = QVBoxLayout()
        layout.addWidget(passphrase_input)
        layout.addWidget(generate_button)
        layout.addWidget(copy_button)

        container = QWidget()
        container.setLayout(layout)
        dialog.layout().addWidget(container)

        generate_button.clicked.connect(
            lambda: self.generate_and_set_passphrase(passphrase_input)
        )
        copy_button.clicked.connect(
            lambda: self.copy_to_clipboard(passphrase_input.text())
        )

        ok_button = dialog.addButton("OK", QMessageBox.AcceptRole)
        cancel_button = dialog.addButton("Annuler", QMessageBox.RejectRole)

        dialog.exec_()

        if dialog.clickedButton() == ok_button:
            return passphrase_input.text(), True
        else:
            return "", False # Demande une passphrase avec option de g√©n√©ration automatique
    def generate_and_set_passphrase(self, input_field):
        new_passphrase = self.generate_secure_passphrase()
        input_field.setText(new_passphrase) # G√©n√®re une passphrase forte et l'affiche
    def copy_to_clipboard(self, text):
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(
            None, "Copi√© !", "Passphrase copi√©e dans le presse-papier."
        ) # Copie la passphrase dans le presse-papier
    def generate_secure_passphrase(self, length=20):
        if length < 8:
            raise ValueError(
                "La longueur minimale de la passphrase doit √™tre de 8 caract√®res."
            )

        while True:
            # Assurer la pr√©sence d'au moins un caract√®re de chaque cat√©gorie requise
            upper = secrets.choice(string.ascii_uppercase)
            digit = secrets.choice(string.digits)
            special = secrets.choice("@$!%*?&")

            # Remplir le reste avec des caract√®res al√©atoires
            characters = string.ascii_letters + string.digits + "@$!%*?&"
            remaining = "".join(secrets.choice(characters) for _ in range(length - 3))

            # M√©langer al√©atoirement pour √©viter un mod√®le fixe
            passphrase = list(upper + digit + special + remaining)
            secrets.SystemRandom().shuffle(passphrase)
            passphrase = "".join(passphrase)

            # V√©rifier que la passphrase respecte les exigences
            if self.is_passphrase_valid(passphrase):
                return passphrase # G√©n√®re une passphrase forte
    def is_passphrase_valid(self, passphrase):
        return bool(
            re.match(
                r"^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$", passphrase
            )
        ) # V√©rifie si la passphrase est valide

# Proc√©dure 1 - Famille de fonctions concernant la r√©cup√©ration et l'int√©gration des valeurs m√©tiers et de leurs besoins
    def convert_level(self, value):
            if value == "Non d√©fini [X]":  # Cas "Non d√©fini" [X]
                return 0
                
            if value == "Elev√©e [H]":  # √âlev√©e
                return 1.5
            elif value == "Moyenne [M]":  # Moyenne
                return 1
            elif value == "Faible [L]":  # Faible
                return 0.5 # Conversion de la m√©trique qualitative s√©mantique CVSS en valeur num√©rique
    def master_function_charger_VM(self): 
        # √âtape 1 : S√©lection du fichier
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "S√©lectionner un fichier Excel",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)",
        )

        if not file_path:
            print("‚ö† Aucun fichier s√©lectionn√©.")
            return

        # Mise √† jour de l'UI
        self.security_input.setText(file_path)
        print(f"üìÇ Fichier s√©lectionn√© : {file_path}")

        # √âtape 2 : Extraction et traitement des valeurs m√©tiers
        try:
            # Charger le fichier Excel
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            results = {}

            # Identifier les colonnes contenant les besoins DIC
            col_dispo = 8   # H - Besoin de disponibilit√© (Dx)
            col_integ = 11  # K - Besoin d‚Äôint√©grit√© (Ix)
            col_confid = 13 # M - Besoin de confidentialit√© (Cx)

            for row in range(3, sheet.max_row + 1):
                valeur_metier = sheet.cell(row=row, column=7).value  # Fonctionnalit√© (colonne G)
                if not valeur_metier:
                    continue

                D = self.convert_level(str(sheet.cell(row=row, column=col_dispo).value))  # Disponibilit√©
                I = self.convert_level(str(sheet.cell(row=row, column=col_integ).value))  # Int√©grit√©
                C = self.convert_level(str(sheet.cell(row=row, column=col_confid).value))  # Confidentialit√©

                results[valeur_metier] = (D, I, C)

            # √âtape 3 : Insertion des donn√©es dans la base de donn√©es
            conn = sqlite3.connect("rbvm.db")
            cur = conn.cursor()

            # Cr√©ation des tables
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS valeurs_metiers(
                    name TEXT PRIMARY KEY,
                    C REAL,
                    I REAL,
                    A REAL      
                );
                """
            )
            conn.commit()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS biens_supports(
                    bs_id TEXT, 
                    cve_id TEXT,
                    bom_serial TEXT,
                    composant_ref TEXT,
                    severity TEXT,
                    score_cvss REAL,
                    attack_vector REAL,
                    attack_complexity REAL,
                    privileges_required REAL,
                    user_interaction REAL,
                    scope TEXT,
                    impact_confidentiality TEXT,
                    impact_integrity TEXT,
                    impact_availability TEXT,
                    exp_score REAL,
                    env_score REAL,
                    KEV TEXT,
                    C_heritage REAL,
                    I_heritage REAL,
                    A_heritage REAL,
                    PRIMARY KEY (bs_id, cve_id)
                );
                """
            )
            conn.commit()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS jointure(
                    num INTEGER PRIMARY KEY AUTOINCREMENT,
                    bs_id TEXT NOT NULL,
                    vm_id TEXT,
                    FOREIGN KEY (vm_id) REFERENCES valeurs_metiers(name)
                );
                """
            )
            conn.commit()

            # Insertion des donn√©es dans la table `valeurs_metiers`
            for valeur_metier, (D, I, C) in results.items():
                cur.execute(
                    """
                    INSERT INTO valeurs_metiers (name, A, C, I)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(name) DO UPDATE SET A=excluded.A, C=excluded.C, I=excluded.I;
                    """,
                    (valeur_metier, D, C, I),
                )

            conn.commit()

            print("‚úÖ Donn√©es des valeurs m√©tiers enregistr√©es avec succ√®s.")

        except Exception as e:
            print(f"‚ùå Erreur lors du traitement du fichier : {e}") # S√©lectionne et charge les besoins de s√©curit√© des valeurs m√©tiers (DIC) depuis un fichier Excel

# Proc√©dure 2 - Famille de fonctions pour le KEV
    def browse_kev(self):
        #"""Ouvre une bo√Æte de dialogue pour s√©lectionner un fichier KEV local ou le t√©l√©charger."""
        self.automatic_download.setChecked(False)
        self.local_file_option.setChecked(True)
        QApplication.processEvents()
        if self.automatic_download.isChecked():
            self.download_kev()  # T√©l√©chargement automatique
        elif self.local_file_option.isChecked():
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "S√©lectionner le fichier KEV",
                "",
                "JSON Files (*.json);;CSV Files (*.csv);;All Files (*)",
            )

            if file_path:
                self.kev_input.setText(file_path)  # Mise √† jour de l'interface
                self.load_kev_data(file_path)  # Charger les donn√©es KEV depuis le fichier s√©lectionn√©
        else:
            QMessageBox.warning(self, "Avertissement", "Veuillez s√©lectionner une m√©thode pour r√©cup√©rer le fichier KEV.") # Fonction nominale pour le KEV
    def browse_kev_manual(self):    
        self.local_file_option.click() # üîπ Simule un clic sur "S√©lectionner un fichier local" pour forcer l'activation
        QApplication.processEvents() # üîπ Forcer la mise √† jour de l'UI pour s'assurer que le changement est pris en compte
        file_path, _ = QFileDialog.getOpenFileName( # üîπ Ouvre la bo√Æte de dialogue pour s√©lectionner un fichier
            self,
            "S√©lectionner le fichier KEV",
            "",
            "JSON Files (*.json);;CSV Files (*.csv);;All Files (*)",
        )

        if file_path:
            self.kev_input.setText(file_path)  # Met √† jour le champ texte
            self.load_kev_data(file_path)  # Fonction permettant de mettre √† jour l'UI d√®s que l'utilisateur veut t√©l√©verser manuellement le KEV Catalog
    def download_kev(self):
        """T√©l√©charge automatiquement le fichier KEV depuis CISA et le charge."""
        url = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
        file_path = "known_exploited_vulnerabilities.json"

        try:
            response = requests.get(url, stream=True)
            response.raise_for_status()  # V√©rifie les erreurs HTTP

            with open(file_path, "wb") as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)

            self.kev_input.setText(file_path)  # Mise √† jour de l'interface
            self.load_kev_data(file_path)  # Charger les donn√©es KEV apr√®s le t√©l√©chargement
            QMessageBox.information(self, "Succ√®s", "Le fichier KEV a √©t√© t√©l√©charg√© et charg√© avec succ√®s.")

        except requests.RequestException as e:
            QMessageBox.critical(self, "Erreur", f"√âchec du t√©l√©chargement : {e}")
    def load_kev_data(self, file_path):
        """Charge les donn√©es KEV depuis un fichier JSON."""
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                self.kev_data = json.load(file)
            print("‚úÖ Donn√©es KEV charg√©es avec succ√®s !")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement du fichier KEV : {e}")
    
# Proc√©dure 3 - Famille de fonctions concernant la r√©cup√©ration et le traitement VDR + KEV

    def master_function_charger_VDR(self):
        """S√©lectionne, charge et traite les fichiers VDR. Met √† jour la base de donn√©es et les vues SQL."""
        print("üìÇ Bouton T√©l√©verser cliqu√©, ouverture du s√©lecteur de fichiers...")

        try:
            # üîπ √âtape 1 : S√©lection des fichiers VDR via bo√Æte de dialogue
            file_paths, _ = QFileDialog.getOpenFileNames(
                self, "S√©lectionner un ou plusieurs VDR", "", "JSON Files (*.json);;All Files (*)"
            )
            print(f"üìÇ Fichiers s√©lectionn√©s : {file_paths}")

            if not file_paths:
                QMessageBox.warning(self, "Avertissement", "Aucun fichier VDR s√©lectionn√©.")
                return  

            self.vdr_data_list = []

            for file_path in file_paths:
                with open(file_path, "r", encoding="utf-8") as file:
                    data = json.load(file)
                    self.vdr_data_list.append(data)

            print(f"üìÑ {len(self.vdr_data_list)} fichier(s) VDR charg√©(s) dans vdr_data_list")

            # üîπ V√©rification de la m√©thode KEV
            if not self.kev_data:
                QMessageBox.warning(self, "Avertissement", "Le fichier KEV n'a pas pu √™tre charg√©.")
                return

            # üîπ √âtape 2 : Connexion √† la base de donn√©es
            conn = sqlite3.connect("rbvm.db")
            cur = conn.cursor()
            print(f"üìÇ Connexion √† la base de donn√©es √©tablie.")

            # üîπ √âtape 3 : Traitement des fichiers VDR
            print(f"üîÑ D√©but du traitement des VDR ({len(self.vdr_data_list)} fichiers)")

            for vdr_data in self.vdr_data_list:
                bs_id = vdr_data.get("metadata", {}).get("component", {}).get("name")
                bom_serial = vdr_data.get("serialNumber")

                # üîç V√©rifier si le bs_id est bien extrait
                print(f"üîç bs_id extrait du VDR : {bs_id}, r√©f√©rence SBOM : {bom_serial}")

                if not bs_id or not bom_serial:
                    print("‚ö† VDR invalide, absence de `bs_id` ou `bom_serial`.")
                    continue

                # üîπ Extraction et insertion des vuln√©rabilit√©s (CVE)
                list_vulnerabilities = vdr_data.get("vulnerabilities", [])

                for vulnerability in list_vulnerabilities:
                    cve_id = vulnerability.get("id")
                    composant_ref = vulnerability.get("bom-ref")
                
                    print(f"üîç cve_id extrait du VDR : {cve_id}, composant_ref extrait du VDR : {composant_ref}")
                    if not cve_id or "GHSA" in cve_id:
                        continue

                    # üîπ Filtrer les entr√©es pour ne prendre que celles avec method="CVSSv3"
                    ratings = [rating for rating in vulnerability.get("ratings", []) if rating.get("method") == "CVSSv3"]

                    if not ratings:
                        print(f"‚ö† Ignor√© : {cve_id} car aucune notation CVSSv3 trouv√©e.")
                        continue  # Passer √† la CVE suivante

                    for rating in ratings:
                        score_CVSS = rating.get("score")
                        severity = rating.get("severity")
                        method = rating.get("method")
                        vector = rating.get("vector", "")

                        if not score_CVSS:
                            continue

                        # V√©rification si c'est une CVE connue exploit√©e (KEV)
                        kev = "YES" if cve_id in self.kev_data else "NO"

                        # üîπ Parsing du vecteur CVSS
                        vector_dic = {key: value for key, value in (item.split(":") for item in vector.replace("CVSS:3.0/", "").split("/")) if key in CVSS_NUMERIC_VALUES}

                        # üîπ Conversion des valeurs CVSS en num√©rique
                        attack_vector = CVSS_NUMERIC_VALUES["AV"].get(vector_dic.get("AV", "N"), 0)
                        attack_complexity = CVSS_NUMERIC_VALUES["AC"].get(vector_dic.get("AC", "L"), 0)
                        privileges_required = CVSS_NUMERIC_VALUES["PR"].get(vector_dic.get("PR", "N"), 0)
                        user_interaction = CVSS_NUMERIC_VALUES["UI"].get(vector_dic.get("UI", "N"), 0)
                        scope = CVSS_NUMERIC_VALUES["S"].get(vector_dic.get("S", "U"), 0)
                        impact_confidentiality = CVSS_NUMERIC_VALUES["C"].get(vector_dic.get("C", "N"), 0)
                        impact_integrity = CVSS_NUMERIC_VALUES["I"].get(vector_dic.get("I", "N"), 0)
                        impact_availability = CVSS_NUMERIC_VALUES["A"].get(vector_dic.get("A", "N"), 0)

                        # üîπ Calcul du score d'exploitabilit√©
                        exp_score = round(
                            (
                                8.22
                                * attack_vector
                                * attack_complexity
                                * privileges_required
                                * user_interaction
                            ) / 3.9 * 4, 
                            1
                        )

                        # üîπ Ajout dans la base de donn√©es
                        try:
                            cur.execute(
                                """
                                INSERT INTO biens_supports(
                                    bs_id, cve_id, bom_serial, composant_ref, severity, score_cvss, KEV,
                                    attack_vector, attack_complexity, privileges_required,
                                    user_interaction, scope, impact_confidentiality,
                                    impact_integrity, impact_availability, exp_score
                                )
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ON CONFLICT (bs_id, cve_id) DO UPDATE SET 
                                    severity=excluded.severity,
                                    score_cvss=excluded.score_cvss,
                                    KEV=excluded.KEV,
                                    attack_vector=excluded.attack_vector,
                                    attack_complexity=excluded.attack_complexity,
                                    privileges_required=excluded.privileges_required,
                                    user_interaction=excluded.user_interaction,
                                    scope=excluded.scope,
                                    impact_confidentiality=excluded.impact_confidentiality,
                                    impact_integrity=excluded.impact_integrity,
                                    impact_availability=excluded.impact_availability,
                                    exp_score=excluded.exp_score;
                                """,
                                (
                                    bs_id, cve_id, bom_serial, composant_ref, severity, score_CVSS, kev,
                                    attack_vector, attack_complexity, privileges_required,
                                    user_interaction, scope, impact_confidentiality,
                                    impact_integrity, impact_availability, exp_score
                                ),
                            )
                            conn.commit()
                            print(f"‚úÖ CVE {cve_id} ins√©r√©e pour {bs_id} avec exp_score: {exp_score}")

                        except sqlite3.Error as e:
                            print(f"‚ùå Erreur SQLite lors de l'insertion de la CVE {cve_id} : {e}")

            print("‚úÖ Mise √† jour des biens supports avec les donn√©es des VDR.")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Une erreur s'est produite : {e}")
    def parse_cvss_vector(self, vector_string):
        """Parse la cha√Æne vectorielle CVSS et extrait les m√©triques."""
        mapping = {
            "AV": {"N": 0.85, "A": 0.62, "L": 0.55, "P": 0.2},
            "AC": {"L": 0.77, "H": 0.44},
            "PR": {"N": 0.85, "L": 0.62, "H": 0.27},
            "UI": {"N": 0.85, "R": 0.62},
            "S": {"U": 0, "C": 1},
            "C": {"N": 0, "L": 0.22, "H": 0.56},
            "I": {"N": 0, "L": 0.22, "H": 0.56},
            "A": {"N": 0, "L": 0.22, "H": 0.56},
        }

        default_values = {metric: "N/A" for metric in mapping.keys()}

        if "CVSS" in vector_string:
            parts = vector_string.split("/")
            for part in parts:
                key_value = part.split(":")
                if len(key_value) == 2 and key_value[0] in mapping:
                    default_values[key_value[0]] = mapping[key_value[0]].get(key_value[1], "N/A")

        return default_values
   
# Proc√©dure 4 -  Association valeur m√©tier / bien support
    def calculer_et_mettre_a_jour_scores_CVSS(self):
        """
        R√©cup√®re les donn√©es de la BDD, calcule les scores environnementaux et met √† jour la base.
        """

        # Dictionnaire pour stocker les scores par bs_id et cve_id
        scores_dic = {}

        # R√©cup√©ration des donn√©es depuis la base de donn√©es
        conn = sqlite3.connect("rbvm.db")
        cur = conn.cursor()
        cur.execute(
            """
            SELECT 
                bs_id,
                cve_id,
                C_heritage, 
                I_heritage, 
                A_heritage, 
                impact_confidentiality, 
                impact_integrity, 
                impact_availability, 
                scope, 
                attack_vector, 
                attack_complexity, 
                privileges_required, 
                user_interaction
            FROM 
                biens_supports
            ORDER BY 
                bs_id, cve_id;
            """
        )
        rows = cur.fetchall()

        if len(rows) == 0:
            print("Aucune donn√©e √† traiter.")
            return

        # üîπ Traitement de chaque ligne
        for row in rows:
            (
                bs_id,
                cve_id,
                C_heritage,
                I_heritage,
                A_heritage,
                impact_confidentiality,
                impact_integrity,
                impact_availability,
                scope,
                attack_vector,
                attack_complexity,
                privileges_required,
                user_interaction,
            ) = row

            # üîπ Remplacement des valeurs None par 0 pour √©viter les erreurs
            C_heritage = float(C_heritage) if C_heritage is not None else 0
            I_heritage = float(I_heritage) if I_heritage is not None else 0
            A_heritage = float(A_heritage) if A_heritage is not None else 0
            impact_confidentiality = float(impact_confidentiality) if impact_confidentiality is not None else 0
            impact_integrity = float(impact_integrity) if impact_integrity is not None else 0
            impact_availability = float(impact_availability) if impact_availability is not None else 0
            scope = float(scope) if scope is not None else 0
            attack_vector = float(attack_vector) if attack_vector is not None else 0
            attack_complexity = float(attack_complexity) if attack_complexity is not None else 0
            privileges_required = float(privileges_required) if privileges_required is not None else 0
            user_interaction = float(user_interaction) if user_interaction is not None else 0

            # üîπ Calcul du score environnemental (fusion de la fonction)
            try:
                if scope == 0:  # Unchanged
                    modified_impact = 6.42 * min(
                        1
                        - (1 - C_heritage * impact_confidentiality)
                        * (1 - I_heritage * impact_integrity)
                        * (1 - A_heritage * impact_availability),
                        0.915,
                    )
                elif scope == 1:  # Changed
                    modified_impact = (
                        7.52
                        * (
                            min(
                                1
                                - (1 - C_heritage * impact_confidentiality)
                                * (1 - I_heritage * impact_integrity)
                                * (1 - A_heritage * impact_availability),
                                0.915,
                            )
                            - 0.029
                        )
                        - 3.25
                        * (
                            (
                                (
                                    min(
                                        1
                                        - (1 - C_heritage * impact_confidentiality)
                                        * (1 - I_heritage * impact_integrity)
                                        * (1 - A_heritage * impact_availability),
                                        0.915,
                                    )
                                )
                                * 0.9731
                                - 0.02
                            )
                            ** 13
                        )
                    )

                modified_exploitability = (
                    8.22
                    * attack_vector
                    * attack_complexity
                    * privileges_required
                    * user_interaction
                )

                if scope == 0:  # Unchanged
                    env_score = round(min(modified_impact + modified_exploitability, 10), 2)
                elif scope == 1:  # Changed
                    env_score = round(
                        min(1.08 * (modified_impact + modified_exploitability), 10), 2
                    )

                # Stockage des scores
                if bs_id not in scores_dic:
                    scores_dic[bs_id] = {}

                scores_dic[bs_id][cve_id] = env_score

            except Exception as e:
                print(f"‚ùå Erreur lors du calcul pour {bs_id}, {cve_id}: {e}")

        # üîπ Mise √† jour de la base de donn√©es
        for bs_id, cve_scores in scores_dic.items():
            for cve_id, env_score in cve_scores.items():
                cur.execute(
                    """
                    UPDATE biens_supports
                    SET env_score = ?
                    WHERE bs_id = ? AND cve_id = ?;
                """,
                    (env_score, bs_id, cve_id),
                )
        conn.commit()

        print("‚úÖ Mise √† jour des scores environnementaux termin√©e.")
        conn.close()
    def process_matrix_data(self):
        """Traitement de la matrice BS-VM et validation avec la base de donn√©es."""

        # √âtape 1 : S√©lection du fichier via l'interface
        self.browse_matrix()

        # V√©rifier si un fichier a bien √©t√© s√©lectionn√©
        file_path = self.matrix_input.text().strip()
        if not file_path:
            print("‚ùå Aucun fichier t√©l√©vers√© √† l'√©tape 2.")
            return

        # √âtape 2 : Extraction des donn√©es de la matrice
        wb = openpyxl.load_workbook(file_path)
        feuille = wb.active

        vm_names = [str(feuille.cell(row=1, column=col).value).strip() for col in range(2, feuille.max_column + 1)
            if feuille.cell(row=1, column=col).value is not None]
        bs_names = [str(feuille.cell(row=row, column=1).value).strip() for row in range(2, feuille.max_row + 1)
            if feuille.cell(row=row, column=1).value is not None]

        results = {}

        # Analyse de la matrice BS-VM
        for col_idx, vm_name in enumerate(vm_names, start=2):
            if not vm_name:
                continue

            for row_idx, bs_name in enumerate(bs_names, start=2):
                cell_value = feuille.cell(row=row_idx, column=col_idx).value

                if cell_value and str(cell_value).strip().lower() == "oui":
                    if vm_name not in results:
                        results[vm_name] = []
                    results[vm_name].append(bs_name)

        print("‚úÖ Matrice BS-VM trait√©e avec succ√®s :", results)

        # √âtape 2B : V√©rification des biens supports et valeurs m√©tiers dans la base
        conn = sqlite3.connect("rbvm.db")
        cur = conn.cursor()

        # üîπ V√©rification des biens supports
        cur.execute("SELECT bs_id FROM biens_supports ORDER BY bs_id;")
        bs_from_db = {row[0] for row in cur.fetchall()}

        missing_bs_in_db = set(bs_names) - bs_from_db
        missing_bs_in_excel = bs_from_db - set(bs_names)

        # üîπ V√©rification des valeurs m√©tiers
        cur.execute("SELECT name FROM valeurs_metiers ORDER BY name;")
        vm_from_db = {row[0] for row in cur.fetchall()}

        missing_vm_in_db = set(vm_names) - vm_from_db
        missing_vm_in_excel = vm_from_db - set(vm_names)

        # ‚ö† Bloquer si des biens supports du fichier Excel ne sont pas en base
        if missing_bs_in_db:
            error_message_bs = (
                "‚ö† ERREUR : Certains biens supports pr√©sents dans le fichier Excel ne sont pas en base de donn√©es !\n\n"
                "Veuillez charger les VDR correspondants avant de continuer.\n\n"
                "üõë Biens supports manquants en base :\n"
                + "\n".join(sorted(str(bs) for bs in missing_bs_in_db if bs is not None))
            )
            print(error_message_bs)
            QMessageBox.critical(self, "Erreur Biens Supports", error_message_bs)
            conn.close()
            return  # Stopper l'ex√©cution ici !

        # ‚ö† Bloquer si des valeurs m√©tiers du fichier Excel ne sont pas en base
        if missing_vm_in_db:
            error_message_vm = (
                "‚ö† ERREUR : Certaines valeurs m√©tiers pr√©sentes dans le fichier Excel ne sont pas en base de donn√©es !\n\n"
                "Veuillez charger le premier fichier Excel des valeurs m√©tiers avant de continuer.\n\n"
                "üõë Valeurs m√©tiers manquantes en base :\n"
                + "\n".join(sorted(str(vm) for vm in missing_vm_in_db if vm is not None))
            )
            print(error_message_vm)
            QMessageBox.critical(self, "Erreur Valeurs M√©tiers", error_message_vm)
            conn.close()
            return  # Stopper l'ex√©cution ici !

        # üõà Information : Afficher si des biens supports existent en base mais ne sont pas dans Excel
        if missing_bs_in_excel:
            print(f"‚ö† Information : Certains biens supports existent en base mais ne sont pas r√©f√©renc√©s dans l'Excel : {missing_bs_in_excel}")

        # üõà Information : Afficher si des valeurs m√©tiers existent en base mais ne sont pas dans Excel
        if missing_vm_in_excel:
            print(f"‚ö† Information : Certaines valeurs m√©tiers existent en base mais ne sont pas r√©f√©renc√©es dans l'Excel : {missing_vm_in_excel}")



        # √âtape 3 : Ins√©rer les relations dans la table `jointure`
        for vm_id, bs_ids in results.items():
            for bs_id in bs_ids:
                cur.execute(
                    """
                    INSERT INTO jointure (bs_id, vm_id)
                    VALUES (?, ?);
                    """,
                    (bs_id, vm_id),
                )
        conn.commit()

        print("‚úÖ Relations BS-VM enregistr√©es avec succ√®s.")

        # √âtape 4 : Mise √† jour des valeurs h√©rit√©es C, I, A
        cur.execute(
            """
            UPDATE biens_supports
            SET 
                C_heritage = (
                    SELECT MAX(vm.C)
                    FROM jointure j
                    JOIN valeurs_metiers vm ON j.vm_id = vm.name
                    WHERE j.bs_id = biens_supports.bs_id
                ),
                I_heritage = (
                    SELECT MAX(vm.I)
                    FROM jointure j
                    JOIN valeurs_metiers vm ON j.vm_id = vm.name
                    WHERE j.bs_id = biens_supports.bs_id
                ),
                A_heritage = (
                    SELECT MAX(vm.A)
                    FROM jointure j
                    JOIN valeurs_metiers vm ON j.vm_id = vm.name
                    WHERE j.bs_id = biens_supports.bs_id
                );
            """
        )
        conn.commit()

        print("‚úÖ Mise √† jour des valeurs h√©rit√©es C, I, A termin√©e.")
        
        # Cr√©ation des vues 

        # Suppression des vues existantes
        cur.execute("DROP VIEW IF EXISTS impact_confidentiality;")
        cur.execute("DROP VIEW IF EXISTS impact_integrity;")
        cur.execute("DROP VIEW IF EXISTS impact_availability;")

        # Cr√©ation des nouvelles vues mises √† jour
        cur.execute("""
            CREATE VIEW impact_confidentiality AS
            SELECT * FROM biens_supports WHERE impact_confidentiality != 0;
        """)

        cur.execute("""
            CREATE VIEW impact_integrity AS
            SELECT * FROM biens_supports WHERE impact_integrity != 0;
        """)

        cur.execute("""
            CREATE VIEW impact_availability AS
            SELECT * FROM biens_supports WHERE impact_availability != 0;
        """)

        conn.commit()  # Enregistre les modifications en base
        
        # calcul score environnemental CVSS

        self.calculer_et_mettre_a_jour_scores_CVSS()  # Calcul et mise √† jour des scores CVSS
        
        conn.close()

# Proc√©dure 5 - Famille de fonctions concernant la g√©n√©ration des repr√©sentations des risques

    def definitionPx(self, p2, p3, p4, p5, scoreEnv, scoreExp):
        """
        Trie les scores d'exploitabilit√© dans les cat√©gories P2 √† P5 selon l'env_score.
        Si une cat√©gorie ne contient qu'une seule valeur, elle est dupliqu√©e pour garantir 
        un affichage correct sur la boxplot.

        Args:
            p2, p3, p4, p5 : Listes contenant les scores tri√©s selon l'env_score.
            scoreEnv (float) : Score environnemental √† classer.
            scoreExp (float) : Score d'exploitabilit√© associ√©.

        Returns:
            p2, p3, p4, p5 : Listes mises √† jour.
        """
    
        if (9.0 <= scoreEnv <= 10.0):
            p2.append(scoreExp)
        elif (7.0 <= scoreEnv <= 8.9):
            p3.append(scoreExp)
        elif (4.0 <= scoreEnv <= 6.9):
            p4.append(scoreExp)
        elif (0.1 <= scoreEnv <= 3.9):
            p5.append(scoreExp)
    
        return p2, p3, p4, p5


    def triAffichageVOR(
        p1,
        p2,
        p3,
        p4,
        p5,
        p1Vert,
        p1Orange,
        p1Rouge,
        p2Vert,
        p2Orange,
        p2Rouge,
        p3Vert,
        p3Orange,
        p4Vert,
        p4Orange,
        p5Vert,
    ):
        for score in p1:
            if 0 <= score <= 0.4:
                p1Vert += 1
            elif 0.5 <= score <= 1.4:
                p1Orange += 1
            else:
                p1Rouge += 1

        for score in p2:
            if 0 <= score <= 0.4:
                p2Vert += 1
            elif 0.5 <= score <= 2.4:
                p2Orange += 1
            else:
                p2Rouge += 1

        for score in p3:
            if 0 <= score <= 1.4:
                p3Vert += 1
            else:
                p3Orange += 1

        for score in p4:
            if 0 <= score <= 2.4:
                p4Vert += 1
            else:
                p4Orange += 1

        for score in p5:
            p5Vert += 1

        return (
            p1Vert,
            p1Orange,
            p1Rouge,
            p2Vert,
            p2Orange,
            p2Rouge,
            p3Vert,
            p3Orange,
            p4Vert,
            p4Orange,
            p5Vert,
        )  # Fonction permettant de trier le nombre de CVE en vert, orange et rouge (VOR) en fonction de chaque P(x)

    def generate_boxplot(self, entity_id, impact, p1, p2, p3, p4, p5, folder_path):
        """
        G√©n√®re un boxplot align√© avec l'image de fond et respecte l'ordre P1 (haut) -> P5 (bas).
        """
        # üîπ Inverser l'ordre des donn√©es pour que P1 soit en haut et P5 en bas
        data = [p1, p2, p3, p4, p5]
        labels = ["P1", "P2", "P3", "P4", "P5"]

        now = datetime.now()
        date = now.strftime("%d.%m.%Y")
        name = f"{entity_id}-{impact}-{date}"

        # URL de l'image de fond
        url = "https://raw.githubusercontent.com/aymericscientist/RVBM_TOOL_LITE/e503bf55ab8210858b010477e419ad3aa2585ae6/fond.png"

        try:
            # T√©l√©charger l'image depuis l'URL
            response = requests.get(url)
            response.raise_for_status()  # V√©rifie si la requ√™te est r√©ussie

            # Charger l'image depuis la m√©moire sans l'enregistrer sur disque
            img = mpimg.imread(BytesIO(response.content), format='png')

            # Cr√©ation du graphique
            fig, ax = plt.subplots(figsize=(10, 5))

            # üîπ Ajustement de l‚Äôimage pour qu‚Äôelle corresponde √† P1 en haut et P5 en bas
            ax.imshow(img, aspect='auto', extent=[0, 4, 5.5, 0.5], alpha=1, zorder=0)

            # üîπ Tracer la bo√Æte √† moustaches par-dessus
            boxprops = dict(facecolor="blue", alpha=0.6)  # Par d√©faut : bleu semi-transparent
            medianprops = dict(color="black", linewidth=1.5)

            # Mise en avant des boxplots o√π exp_score = 4
            if all(np.median(group) == 4 for group in data if len(group) > 0):
                medianprops = dict(color="black", linewidth=3)

            ax.boxplot(data, vert=False, patch_artist=True, showfliers=True, labels=labels, 
                       boxprops=boxprops, medianprops=medianprops, zorder=1, whis=[0, 100])

            # üîπ Ajouter titre et labels
            ax.set_title(f"{entity_id} - {impact.capitalize()}")
            ax.set_xlabel("Score d'exploitabilit√©")

            # üîπ Ajuster l'axe X (score exploitabilit√©) de 0 √† 4 avec des intervalles de 0.5
            ax.set_xlim(0, 4)
            ax.set_xticks([i * 0.5 for i in range(9)])  # De 0 √† 4 avec un pas de 0.5

            # üîπ Ajuster l'axe Y pour afficher P1 en haut et P5 en bas
            ax.set_ylim(5.5, 0.5)  # üîπ Correction ici pour inverser l'affichage
            ax.set_yticks([1, 2, 3, 4, 5])
            ax.set_yticklabels(["P1", "P2", "P3", "P4", "P5"])  # üîπ Ordre corrig√©

            # üîπ Sauvegarde de l'image avec la boxplot
            plt.savefig(f"{folder_path}/{name}.png", format="png", dpi=300)
            plt.close(fig)

            print(f"‚úÖ Boxplot align√©e avec l'image de fond : {folder_path}/{name}.png")

        except requests.RequestException as e:
            print(f"‚ùå Erreur lors du t√©l√©chargement de l'image de fond : {e}")







    def boite(self, boite_path):
        """
        G√©n√®re des bo√Ætes √† moustaches pour chaque bs_id √† partir des vues impact_availability, 
        impact_confidentiality et impact_integrity en respectant les crit√®res stricts de classification.
        """
        cur = self.cur  # Utilisation du curseur SQLite d√©j√† initialis√©
        liste_impacts = ["availability", "confidentiality", "integrity"]
        vues_impact = {
            "availability": "impact_availability",
            "confidentiality": "impact_confidentiality",
            "integrity": "impact_integrity",
        }

        for impact in liste_impacts:
            vue = vues_impact[impact]

            # üîπ R√©cup√©ration des `bs_id` distincts pour le type d'impact donn√©
            cur.execute(f"SELECT DISTINCT bs_id FROM {vue};")
            bs_list = cur.fetchall()

            for bs in bs_list:
                bs_id = bs[0]

                # üîπ R√©cup√©ration des valeurs exp_score, env_score et kev pour chaque `bs_id` depuis la vue correspondante
                cur.execute(
                    f"""
                    SELECT exp_score, env_score, kev 
                    FROM {vue} 
                    WHERE bs_id = ?;
                    """,
                    (bs_id,),
                )
                data = cur.fetchall()

                # üîπ Initialisation des listes pour stocker les scores class√©s
                p1, p2, p3, p4, p5 = [], [], [], [], []

                for exp_score, env_score, kev in data:
                    if kev == "YES":
                        p1.append(exp_score)
                    elif 9.00 <= env_score <= 10.0:
                        p2.append(exp_score)
                    elif 7.00 <= env_score <= 8.9:
                        p3.append(exp_score)
                    elif 4.00 <= env_score <= 6.9:
                        p4.append(exp_score)
                    elif 0.1 <= env_score <= 3.9:
                        p5.append(exp_score)

                # üîπ G√©n√©ration de la bo√Æte √† moustaches uniquement si des donn√©es existent
                if any([p1, p2, p3, p4, p5]):
                    self.generate_boxplot(bs_id, impact, p1, p2, p3, p4, p5, boite_path)




    def boite_vm(self, vm_id, folder_path):
        liste_dia = ["confidentiality", "integrity", "availability"]
        cur.execute(
            """
            SELECT j.bs_id
            FROM jointure j
            WHERE j.vm_id = ?;
        """,
            (vm_id,),
        )
        rows = cur.fetchall()

        p1t, p2t, p3t, p4t, p5t = {}, {}, {}, {}, {}
        for impact in liste_dia:
            p1t[impact] = []
            p2t[impact] = []
            p3t[impact] = []
            p4t[impact] = []
            p5t[impact] = []

        for row in rows:
            bs_id = row[0]
            for impact in liste_dia:
                impact_key = f"{bs_id}-{impact}"
                if impact_key in dicoListePx:
                    p5t[impact].extend(dicoListePx[impact_key][0])
                    p4t[impact].extend(dicoListePx[impact_key][1])
                    p3t[impact].extend(dicoListePx[impact_key][2])
                    p2t[impact].extend(dicoListePx[impact_key][3])
                    p1t[impact].extend(dicoListePx[impact_key][4])

        for impact in liste_dia:
            dicoGlobal[impact]["p5"].extend(p5t[impact])
            dicoGlobal[impact]["p4"].extend(p4t[impact])
            dicoGlobal[impact]["p3"].extend(p3t[impact])
            dicoGlobal[impact]["p2"].extend(p2t[impact])
            dicoGlobal[impact]["p1"].extend(p1t[impact])

            self.generate_boxplot(vm_id, impact, p1t[impact], p2t[impact], p3t[impact], p4t[impact], p5t[impact], folder_path) # Cr√©ation des bo√Ætes √† moustaches pour les valeurs m√©tiers
    def boite_vm_globale(self, folder_path):
        liste_dia = ["confidentiality", "integrity", "availability"]

        for impact in liste_dia:
            p1 = dicoGlobal[impact]["p1"]
            p2 = dicoGlobal[impact]["p2"]
            p3 = dicoGlobal[impact]["p3"]
            p4 = dicoGlobal[impact]["p4"]
            p5 = dicoGlobal[impact]["p5"]

            self.generate_boxplot("Meta_VM", impact, p1, p2, p3, p4, p5, folder_VM_META) # Cr√©ation des bo√Ætes √† moustaches pour la repr√©sentation globale des valeurs m√©tiers


    def start_conversion_MOE(self):
        print("D√©but de la g√©n√©ration des repr√©sentations MOE")
        boite_path = filedialog.askdirectory()
        if not boite_path:
            print("No folder selected.")
            return
        
        print("üìä V√©rification du contenu de la base SQLite")
        cur.execute("SELECT * FROM biens_supports LIMIT 5;")
        rows = cur.fetchall()
        if not rows:
            print("‚ö† Aucun bien support enregistr√© dans la base de donn√©es.")
            return

        print("üìä V√©rification des relations jointure")
        cur.execute("SELECT * FROM jointure LIMIT 5;")
        jointure_rows = cur.fetchall()
        if not jointure_rows:
            print("‚ö† Aucune relation BS-VM enregistr√©e dans `jointure`.")
            return

        print("üìä V√©rification des valeurs h√©rit√©es C, I, A")
        cur.execute("SELECT bs_id, C_heritage, I_heritage, A_heritage FROM biens_supports LIMIT 5;")
        heritage_rows = cur.fetchall()
        for row in heritage_rows:
            print(f"  BS: {row[0]}, C: {row[1]}, I: {row[2]}, A: {row[3]}")

        print("üìä V√©rification avant ex√©cution de `boite()`")
        print(f"üìä Contenu actuel de `dicoListePx` avant ex√©cution : {dicoListePx}")

        # Appel de la fonction pour g√©n√©rer les bo√Ætes √† moustaches
        self.boite(boite_path)

        print("üìä Contenu apr√®s ex√©cution de `boite()` :", dicoListePx)

        subfolder_VERT = os.path.join(boite_path, "03_VERT")
        subfolder_ORANGE = os.path.join(boite_path, "02_ORANGE")
        subfolder_ROUGE = os.path.join(boite_path, "01_ROUGE")
        folder_BS_VERT = subfolder_VERT
        folder_BS_ORANGE = subfolder_ORANGE
        folder_BS_ROUGE = subfolder_ROUGE
        if not os.path.exists(subfolder_VERT):
            os.makedirs(subfolder_VERT, exist_ok=True)
        if not os.path.exists(subfolder_ORANGE):
            os.makedirs(subfolder_ORANGE, exist_ok=True)
        if not os.path.exists(subfolder_ROUGE):
            os.makedirs(subfolder_ROUGE, exist_ok=True)
        self.boite(boite_path)
    def start_conversion_MOA(self):
        print("‚úÖ start_conversion_MOA appel√©")
        print("üìä Contenu actuel de dicoListePx :", dicoListePx)  # V√©rification
        if not dicoListePx:
            print("‚ùå dicoListePx est vide ou non d√©fini !")
            return
        folder_path = filedialog.askdirectory()
        if not folder_path:
            print("No folder selected.")
            return
        
        # Cr√©er les sous-dossiers pour les diff√©rentes cat√©gories
        subfolder_ROUGE = os.path.join(folder_path, "01_ROUGE")
        subfolder_ORANGE = os.path.join(folder_path, "02_ORANGE")
        subfolder_VERT = os.path.join(folder_path, "03_VERT")
        subfolder_VM_META = os.path.join(folder_path, "04_Meta_repr√©sentation_des_VM")
        folder_VM_VERT = subfolder_VERT
        folder_VM_ORANGE = subfolder_ORANGE
        folder_VM_ROUGE = subfolder_ROUGE
        folder_VM_META = subfolder_VM_META
        if not os.path.exists(subfolder_VERT):
            os.makedirs(subfolder_VERT, exist_ok=True)
        if not os.path.exists(subfolder_ORANGE):
            os.makedirs(subfolder_ORANGE, exist_ok=True)
        if not os.path.exists(subfolder_ROUGE):
            os.makedirs(subfolder_ROUGE, exist_ok=True)
        if not os.path.exists(subfolder_VM_META):
            os.makedirs(subfolder_VM_META, exist_ok=True)
        cur.execute("SELECT DISTINCT vm_id FROM jointure;")
        v = cur.fetchall()
        for vid in v:
            vid = vid[0]
            self.boite_vm(vid, folder_path)
        self.boite_vm_globale(folder_path)

# Famille de fonctions concernant l'IHM
    def initUI(self):
        """Initialise l'interface utilisateur"""
        central_widget = QWidget()
        main_layout = QVBoxLayout()

        # Titre principal
        title = QLabel("Risk Based Vulnerability Management (RBVM) Tool", self)
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #444;")

        # 1√®re √©tape - Int√©grer les besoins de s√©curit√© et s√ªret√© des valeurs m√©tiers
        security_group = QGroupBox(
            "1√®re √©tape [MOA] : Charger (excel) les valeurs m√©tiers ainsi que leurs besoins de s√©curit√© et s√ªret√© [template_prerequis DIC.xlsx]"
        )
        security_layout = QHBoxLayout()
        self.security_input = QLineEdit(self)
        security_browse = QPushButton("T√©l√©verser", self)
        security_browse.clicked.connect(self.master_function_charger_VM)
        security_layout.addWidget(self.security_input)
        security_layout.addWidget(security_browse)
        security_group.setLayout(security_layout)

        # 2√®me √©tape - Charger le fichier KEV Catalog
        kev_group = QGroupBox(
            "2√®me √©tape [MOE] : Charger le fichier Known Exploited Vulnerabilities (KEV) Catalog du Cybersecurity & Infrastructure Security Agency (CISA)"
        )
        kev_layout = QVBoxLayout()

        self.automatic_download = QRadioButton(
            "T√©l√©charger automatiquement depuis CISA [par d√©faut]", self
        )
        self.local_file_option = QRadioButton(
            "S√©lectionner un fichier local (JSON/CSV)", self
        )
        self.automatic_download.setChecked(
            True
        )  # Option par d√©faut : T√©l√©charger automatiquement depuis CISA

        kev_file_layout = QHBoxLayout()
        self.kev_input = QLineEdit(self)
        kev_browse = QPushButton("T√©l√©verser", self)
        kev_browse.clicked.connect(self.browse_kev_manual)
        kev_file_layout.addWidget(self.kev_input)
        kev_file_layout.addWidget(kev_browse)

        kev_layout.addWidget(self.automatic_download)
        kev_layout.addWidget(self.local_file_option)
        kev_layout.addLayout(kev_file_layout)
        kev_group.setLayout(kev_layout)

        # 3√®me √©tape - Charger VDR
        vdr_group = QGroupBox(
            "3√®me √©tape [MOE] : Charger tous les Vulnerability Disclosure Report (VDR) concernant l'exhaustivit√© des biens supports"
        )
        vdr_layout = QHBoxLayout()
        self.vdr_input = QLineEdit(self)
        vdr_browse = QPushButton("T√©l√©verser", self)
        vdr_browse.clicked.connect(self.master_function_charger_VDR)
        vdr_layout.addWidget(self.vdr_input)
        vdr_layout.addWidget(vdr_browse)
        vdr_group.setLayout(vdr_layout)

        # 4√®me √©tape - Int√©grer la matrice Bien Support - Valeur M√©tier
        matrix_group = QGroupBox(
            "4√®me √©tape [MOE] : Charger la matrice (excel) associant les biens supports aux valeurs m√©tiers [template_matrice_vm_bs.xlsx]"
        )
        matrix_layout = QHBoxLayout()
        self.matrix_input = QLineEdit(self)
        matrix_browse = QPushButton("T√©l√©verser", self)
        matrix_browse.clicked.connect(self.process_matrix_data)
        matrix_layout.addWidget(self.matrix_input)
        matrix_layout.addWidget(matrix_browse)
        matrix_group.setLayout(matrix_layout)

        # Ajouter un layout horizontal pour les boutons
        buttons_layout = QHBoxLayout()

        # Bouton g√©n√©ration repr√©sentation des risques concernant les biens supports (RSSI)
        self.convert_button_MOE = QPushButton(
            "G√©n√©rer les repr√©sentations concernant les risques li√©s aux biens supports (MOE)",
            self,
        )
        self.convert_button_MOE.setStyleSheet(
            "background-color: #28a745; color: white; font-size: 12px; padding: 8px;"
        )
        self.convert_button_MOE.clicked.connect(self.start_conversion_MOE)

        # Bouton g√©n√©ration repr√©sentation des risques concernant les valeurs m√©tiers (CSN)
        self.convert_button_MOA = QPushButton(
            "G√©n√©rer les repr√©sentations concernant les risques li√©s aux valeurs m√©tiers (MOA)",
            self,
        )
        self.convert_button_MOA.setStyleSheet(
            "background-color: #28a745; color: white; font-size: 12px; padding: 8px;"
        )
        self.convert_button_MOA.clicked.connect(self.start_conversion_MOA)
        
        # Ajouter les boutons au layout horizontal
        buttons_layout.addWidget(self.convert_button_MOE)
        buttons_layout.addWidget(self.convert_button_MOA)

        # Ajout au layout principal
        main_layout.addWidget(title)
        main_layout.addWidget(security_group)  # Charger besoins de s√©curit√© et s√ªret√©
        main_layout.addWidget(kev_group)  # Charger KEV
        main_layout.addWidget(vdr_group)  # S√©lectionner les VDR
        main_layout.addWidget(matrix_group)  # Charger association VM & BS
        main_layout.addLayout(buttons_layout)

        # Appliquer le layout principal √† la fen√™tre
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Appliquer le style CSS
        self.setStyleSheet(
            """
            QMainWindow {
                background-color: #f4f4f4;
            }
            QLabel {
                font-size: 14px;
                color: #333;
            }
            QPushButton {
                background-color: #0078D7;
                color: white;
                font-size: 14px;
                padding: 6px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #005BB5;
            }
            QLineEdit {
                border: 1px solid #ccc;
                padding: 5px;
                border-radius: 5px;
                background-color: white;
            }
            QProgressBar {
                border: 1px solid #bbb;
                padding: 3px;
                border-radius: 5px;
                background-color: #fff;
                text-align: center;
            }
        """
        )
    def browse_kev(self):
        """Ouvre une bo√Æte de dialogue pour s√©lectionner un fichier KEV local (JSON/CSV)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "S√©lectionner un fichier KEV",
            "",
            "JSON/CSV Files (*.json *.csv);;All Files (*)",
        )
        if file_path:
            self.kev_input.setText(file_path)  # Int√©grer le KEV Catalog
    def browse_matrix(self):
        """Ouvre une bo√Æte de dialogue pour s√©lectionner un fichier Excel contenant la matrice BS-VM"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "S√©lectionner un fichier Excel",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)",
        )
        if file_path:
            self.matrix_input.setText(
                file_path
            )  # Int√©grer la matrice valeur m√©tier - bien support
    def browse_security(self):
        """Ouvre une bo√Æte de dialogue pour s√©lectionner un fichier Excel contenant les besoins de s√©curit√©"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "S√©lectionner un fichier Excel",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)",
        )
        if file_path:
            self.security_input.setText(
                file_path
            )  # Int√©grer les besoins de s√©curit√© et s√ªret√© des valeurs m√©tiers

if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    fenetre = RBVMTool()
    fenetre.show()
    root = Tk()
    root.withdraw()

    ## ==== pr√©requis BDD du programme ==== ##
    conn = sqlite3.connect("rbvm.db")  # Connexion √† la base de donn√©es SQLite chiffr√©e avec SQLCipher
    conn.execute(f'PRAGMA passphrase = "{fenetre.passphrase}";')  # Appliquer la cl√© PRAGMA chiffr√©e √† la connexion SQLite
    conn.execute("PRAGMA foreign_passphrase = ON;")  # Appliquer la cl√© PRAGMA chiffr√©e √† la connexion SQLite
    cur = conn.cursor()  # Cr√©ation du curseur



    sys.exit(app.exec_())  #  Lancement de l'application







##### CLASSES  #####
class CVE:
    def __init__(
        self,
        id,
        composant_ref,
        description,
        severity,
        score_CVSS,
        attack_vector,
        attack_complexity,
        privileges_required,
        user_interaction,
        scope,
        confidentiality,
        integrity,
        availability,
        exp_score,
        env_score,
        kev,
    ):
        self.id = id
        self.composant_ref = composant_ref
        self.description = description
        self.severity = severity
        self.score_CVSS = score_CVSS
        self.attack_vector = attack_vector
        self.attack_complexity = attack_complexity
        self.privileges_required = privileges_required
        self.user_interaction = user_interaction
        self.scope = scope
        self.confidentiality = confidentiality
        self.integrity = integrity
        self.availability = availability
        self.exp_score = exp_score
        self.env_score = env_score
        self.kev = kev

    def __repr__(self):
        return f"ID: {self.id} \nBom-ref: {self.composant_ref} \nDescription: {self.description} \nSeverity: {self.severity} \nScore CVSS: {self.score_CVSS} \nAV: {self.attack_vector} \nAC: {self.attack_complexity} \nPR: {self.privileges_required} \nUI: {self.user_interaction} \nS: {self.scope} \nC: {self.confidentiality} \nI: {self.integrity} \nA: {self.availability} \nScore Exp: {self.exp_score} \nScore Env: {self.env_score} \nKeV: {self.kev}\n\n"  # Classe permettant de d√©finir les CVE
class CVE_others:
    def __init__(self, id, composant_ref, description, severity):
        self.id = id
        self.composant_ref = composant_ref
        self.description = description
        self.severity = severity

    def __repr__(self):
        return f"ID: {self.id} \nBom-ref: {self.composant_ref} \nDescription: {self.description} \nSeverity: {self.severity}\n\n"


##### FONCTIONS #####

# Famille de fonctions pour parser les variables environnementales √† partir du VDR (CVSS X)
def var_environnementales_CVSSv3(svector):
    attack_vector = svector[12]
    attack_complexity = svector[17]
    privileges_required = svector[22]
    user_interaction = svector[27]
    scope = svector[31]
    confidentiality = svector[35]
    integrity = svector[39]
    availability = svector[43]
    return (
        attack_vector,
        attack_complexity,
        privileges_required,
        user_interaction,
        scope,
        confidentiality,
        integrity,
        availability,
    )  # Fonction permettant de parser les variables environnementales √† partir du "vector" du VDR (CVSS3.0)
def calcul_score_exploitabilit√©(
    attack_vector, attack_complexity, privileges_required, user_interaction
):

    exp_score = (
        (
            8.22
            * attack_vector
            * attack_complexity
            * privileges_required
            * user_interaction
        )
        / 3.9
        * 4
    )
    return round(
        exp_score, 1
    )  # Fonction permettant de calculer le score d'exploitabilit√© CVSS 3.1

def calcul_score_environnemental(
    disponibiliteVM,
    integriteVM,
    confidentialiteVM,
    availability,
    integrity,
    confidentiality,
    scope,
    attack_vector,
    attack_complexity,
    privileges_required,
    user_interaction,
):
    try:
        disponibiliteVM = float(disponibiliteVM)
        integriteVM = float(integriteVM)
        confidentialiteVM = float(confidentialiteVM)
        availability = float(availability)
        integrity = float(integrity)
        confidentiality = float(confidentiality)
        attack_vector = float(attack_vector)
        attack_complexity = float(attack_complexity)
        privileges_required = float(privileges_required)
        user_interaction = float(user_interaction)
    except ValueError as e:
        print(f"Erreur de conversion: {e}")
        return None  # ou une valeur par d√©faut

    # D√©finition de la valeur scope
    if "U" in scope:
        modified_impact = 6.42 * min(
            1
            - (1 - disponibiliteVM * availability)
            * (1 - integriteVM * integrity)
            * (1 - confidentialiteVM * confidentiality),
            0.915,
        )
    elif "C" in scope:
        modified_impact = 7.52 * (
            min(
                1
                - (1 - disponibiliteVM * availability)
                * (1 - integriteVM * integrity)
                * (1 - confidentialiteVM * confidentiality),
                0.915,
            )
            - 0.029
        ) - 3.25 * (
            (
                (
                    min(
                        1
                        - (1 - disponibiliteVM * availability)
                        * (1 - integriteVM * integrity)
                        * (1 - confidentialiteVM * confidentiality),
                        0.915,
                    )
                )
                * 0.9731
                - 0.02
            )
            ** 13
        )

    modified_exploitability = (
        8.22
        * attack_vector
        * attack_complexity
        * privileges_required
        * user_interaction
    )

    if "U" in scope:
        env_score = round(min(modified_impact + modified_exploitability, 10), 2)
    elif "C" in scope:
        env_score = round(
            min(1.08 * (modified_impact + modified_exploitability), 10), 2
        )

    return (
        env_score  # Fonction permettant de calculer le score environnemental CVSS 3.1
    )
def option_5_calcul_scores():
    # Dictionnaire pour stocker les scores par bs_id et cve_id
    scores_dic = {}

    # R√©cup√©ration des donn√©es
    cur.execute(
        """
        SELECT 
            bs_id,
            cve_id,
            C_heritage, 
            I_heritage, 
            A_heritage, 
            impact_confidentiality, 
            impact_integrity, 
            impact_availability, 
            scope, 
            attack_vector, 
            attack_complexity, 
            privileges_required, 
            user_interaction 
        FROM 
            biens_supports
        ORDER BY 
            bs_id, cve_id;
    """
    )
    rows = cur.fetchall()

    if len(rows) == 0:
        print("Aucune donn√©e √† traiter.")
        return

    # Traitement de chaque ligne
    for row in rows:
        (
            bs_id,
            cve_id,
            C_heritage,
            I_heritage,
            A_heritage,
            impact_confidentiality,
            impact_integrity,
            impact_availability,
            scope,
            attack_vector,
            attack_complexity,
            privileges_required,
            user_interaction,
        ) = row

        # Conversion des valeurs textuelles en num√©riques
        impact_availability_num = convert_cia_to_numeric(impact_availability)
        impact_integrity_num = convert_cia_to_numeric(impact_integrity)
        impact_confidentiality_num = convert_cia_to_numeric(impact_confidentiality)

        try:
            # Calcul du score environnemental
            env_score = calcul_score_environnemental(
                A_heritage,
                I_heritage,
                C_heritage,
                impact_availability_num,
                impact_integrity_num,
                impact_confidentiality_num,
                scope,
                attack_vector,
                attack_complexity,
                privileges_required,
                user_interaction,
            )

            # Stockage des scores
            if bs_id not in scores_dic:
                scores_dic[bs_id] = {}

            scores_dic[bs_id][cve_id] = env_score

        except Exception as e:
            print(f"Erreur lors du calcul pour {bs_id}, {cve_id}: {e}")

    # Mise √† jour de la base de donn√©es
    for bs_id, cve_scores in scores_dic.items():
        for cve_id, env_score in cve_scores.items():
            cur.execute(
                """
                UPDATE biens_supports
                SET env_score = ?
                WHERE bs_id = ? AND cve_id = ?;
            """,
                (env_score, bs_id, cve_id),
            )
def convert_cia_to_numeric(value):
    # Convertit les valeurs d'impact CIA en valeurs num√©riques
    if value == "N":
        return 0  # None
    elif value == "L":
        return 0.22  # Low
    elif value == "H":
        return 0.56  # High
    return 0  # Fonctions de conversion des valeurs d'impact CIA en valeurs num√©riques


